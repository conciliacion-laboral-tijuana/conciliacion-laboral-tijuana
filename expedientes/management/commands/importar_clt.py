"""
Comando de gestión: importar_clt

Importa el archivo CLT.xlsx (calendario de citas de conciliación) al sistema,
creando Cliente + Expediente con fecha de audiencia para que las citas
aparezcan en el calendario de audiencias.

Estructura esperada del archivo (sin encabezados, fila 2 en adelante):
    A: Nombre del trabajador (cliente)
    B: Empresa / patrón
    C: Fecha de la cita       (ej: "02 DE JULIO DE 2026" o datetime)
    D: Hora de la cita        (ej: 15:01:00 o "N/A")
    E: Salario mensual
    F: Honorario estimado     (30% del salario - 1200, fórmula en el archivo)
    G: Asesor principal       (ej: "PATRICIA/CONY", "Kevin", "DULCE")
    H: Asesor secundario      (opcional)
    I: Ubicación CLT          (CLT OTAY / CLT CJL / CLT PP)
    J, K: Fracciones del honorario (no se importan)
    L, M: Datos extra         (se guardan como nota si existen)

Uso:
    uv run python manage.py importar_clt                    # Importa CLT.xlsx (raíz del proyecto)
    uv run python manage.py importar_clt --archivo ruta/CLT.xlsx
    uv run python manage.py importar_clt --hoja AGOSTO      # Solo una hoja
    uv run python manage.py importar_clt --crear-asesores   # Crea los asesores del archivo como usuarios reales
    uv run python manage.py importar_clt --solo-asesores    # Solo crea los asesores, sin importar citas
    uv run python manage.py importar_clt --borrar-prueba    # Elimina usuarios de prueba (asesor1-15, admin1-4, etc.)
    uv run python manage.py importar_clt --borrar-prueba --crear-asesores --archivo CLT.xlsx
"""
import re
import unicodedata
from datetime import datetime, time as dt_time
from decimal import Decimal, InvalidOperation
from io import StringIO

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand
from django.utils import timezone

from expedientes.models import Cliente, Expediente

# ─── Usuarios de prueba creados por crear_usuarios_prueba ──────────────────
USUARIOS_PRUEBA = (
    ['superadmin'] +
    [f'admin{i}' for i in range(1, 5)] +
    [f'asesor{i}' for i in range(1, 16)] +
    ['finanzas1']
)

MESES_ES = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
    'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
    'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12,
}

# Columna I del archivo -> oficina del Cliente
UBICACION_OFICINA = {
    'OTAY': 'otay',
    'CJL': 'clt',   # CJL (Centro de Justicia Laboral) → CLT
    'CLT': 'clt',
    'PP': 'plaza_patria',
}

PASSWORD_ASESORES = 'Asesor123!'  # Misma convención que crear_usuarios_prueba


def _normalizar(texto):
    """Mayúsculas, sin acentos, espacios compactados."""
    if not texto:
        return ''
    texto = unicodedata.normalize('NFKD', str(texto))
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', texto).strip().upper()


def _asesor_principal(valor):
    """Extrae el nombre del asesor principal de la columna G.

    Maneja "PATRICIA/CONY" -> PATRICIA, "DAMARIS/BRISA" -> DAMARIS,
    "DULCE/3PM+" -> DULCE, "Kevin 3pm+" -> KEVIN, "Nínive" -> NINIVE.
    """
    if not valor:
        return ''
    texto = _normalizar(valor)
    # Tomar la primera parte antes de separadores
    parte = re.split(r'[/,|]', texto)[0]
    # Quitar sufijos tipo "3PM+", "FUERA", "PM"
    parte = re.sub(r'\s*(3PM\+?|PM|FUERA|\+).*$', '', parte)
    return parte.strip()


def _nombre_usuario(nombre_normalizado):
    """Convierte un nombre normalizado a username: minúsculas, sin acentos."""
    return nombre_normalizado.lower()


def _parsear_fecha(valor):
    """Parsea la fecha de la cita: datetime, '02 DE JULIO DE 2026' o 'N/A'."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    texto = _normalizar(valor)
    if not texto or texto in ('N/A', 'NA', 'S/F', 'PENDIENTE'):
        return None
    m = re.match(r'^(\d{1,2})\s+DE\s+([A-ZÑ]+)\s+DE\s+(\d{4})$', texto)
    if m:
        dia, mes, anio = m.groups()
        mes_num = MESES_ES.get(mes)
        if mes_num:
            try:
                return datetime(int(anio), mes_num, int(dia)).date()
            except ValueError:
                return None
    # Fallback: fechas ISO o dd/mm/yyyy
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _parsear_hora(valor):
    """Parsea la hora: datetime.time o 'N/A'."""
    if valor is None:
        return None
    if isinstance(valor, dt_time):
        return valor
    texto = _normalizar(valor)
    if not texto or texto in ('N/A', 'NA'):
        return None
    m = re.match(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?$', texto)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        s = int(m.group(3) or 0)
        try:
            return dt_time(h, mi, s)
        except ValueError:
            return None
    return None


def _decimal(valor):
    if valor is None or valor == '':
        return None
    try:
        return Decimal(str(valor))
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = 'Importa citas del archivo CLT.xlsx como expedientes con audiencia programada'

    def add_arguments(self, parser):
        parser.add_argument('--archivo', default='CLT.xlsx',
                            help='Ruta del archivo xlsx (default: CLT.xlsx en la raíz)')
        parser.add_argument('--hoja', default='',
                            help='Solo importar esta hoja (JULIO, AGOSTO, ...)')
        parser.add_argument('--crear-asesores', action='store_true',
                            help='Crea como usuarios reales (rol asesor) los asesores que aparecen en el archivo')
        parser.add_argument('--solo-asesores', action='store_true',
                            help='Solo crea los asesores del archivo; NO importa citas')
        parser.add_argument('--borrar-prueba', action='store_true',
                            help='Elimina los usuarios de prueba (asesor1-15, admin1-4, superadmin, finanzas1) '
                                 'y sus expedientes, ANTES de importar')

    # ─── Borrado de usuarios de prueba ───────────────────────────────────────

    def _borrar_usuarios_prueba(self):
        test_users = list(User.objects.filter(username__in=USUARIOS_PRUEBA))
        if not test_users:
            self.stdout.write(self.style.WARNING('  No hay usuarios de prueba para borrar.'))
            return

        # Eliminar expedientes de esos usuarios (FK PROTECT no permite borrar user con expedientes)
        exp_count = Expediente.objects.filter(asesor__in=test_users).count()
        if exp_count:
            Expediente.objects.filter(asesor__in=test_users).delete()
            self.stdout.write(self.style.WARNING(f'  -> Eliminados {exp_count} expedientes de prueba'))

        # Eliminar clientes huérfanos (sin expedientes) para no duplicar al re-importar
        from expedientes.models import Cliente as C
        huérfanos = C.objects.filter(expediente__isnull=True).count()
        if huérfanos:
            C.objects.filter(expediente__isnull=True).delete()
            self.stdout.write(self.style.WARNING(f'  -> Eliminados {huérfanos} clientes sin expedientes'))

        users_borrados = 0
        for u in test_users:
            # Si aun tiene referencias (p. ej. movimientos), borrar en cascada
            u.delete()
            users_borrados += 1
        self.stdout.write(self.style.SUCCESS(f'  [OK] Usuarios de prueba eliminados: {users_borrados}'))

    # ─── Creación de asesores reales ─────────────────────────────────────────

    def _crear_asesores(self, nombres):
        """Crea (o reutiliza) un User con rol asesor por cada nombre único del archivo.
        Retorna (creados, existentes)."""
        creados, existentes = [], []
        for nombre in sorted(nombres):
            if not nombre:
                continue
            username = _nombre_usuario(nombre)
            user = User.objects.filter(username=username).first()
            if user:
                existentes.append(username)
                continue
            user = User.objects.create_user(
                username=username,
                email=f'{username}@despacho.mx',
                password=PASSWORD_ASESORES,
                first_name=nombre.title(),
            )
            user.profile.rol = 'asesor'
            user.profile.save()
            creados.append(username)
        if creados:
            self.stdout.write(self.style.SUCCESS(f'  [OK] Asesores creados: {", ".join(creados)} (pass: {PASSWORD_ASESORES})'))
        if existentes:
            self.stdout.write(f'  -> Asesores ya existentes: {", ".join(existentes)}')
        return creados, existentes

    # ─── Importación ─────────────────────────────────────────────────────────

    def _leer_filas(self, ruta, hoja_sel):
        """Lee las filas del xlsx. Retorna lista de dicts."""
        import openpyxl
        wb = openpyxl.load_workbook(ruta, data_only=True)
        hojas = [hoja_sel] if hoja_sel else wb.sheetnames
        filas = []
        for nombre_hoja in hojas:
            if nombre_hoja not in wb.sheetnames:
                self.stdout.write(self.style.ERROR(f'  [X] Hoja "{nombre_hoja}" no existe en el archivo'))
                continue
            ws = wb[nombre_hoja]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue  # fila vacía
                filas.append({
                    'hoja': nombre_hoja,
                    'nombre': str(row[0]).strip(),
                    'empresa': str(row[1]).strip() if row[1] else '',
                    'fecha': _parsear_fecha(row[2]),
                    'hora': _parsear_hora(row[3]),
                    'salario': _decimal(row[4]),
                    'honorario': _decimal(row[5]),
                    'asesor': _asesor_principal(row[6]),
                    'asesor2': _normalizar(row[7]) if row[7] else '',
                    'ubicacion': _normalizar(row[8]) if row[8] else '',
                    'extra': _normalizar(row[11]) if len(row) > 11 and row[11] else '',
                })
        return filas

    def _importar_filas(self, filas):
        """Crea/actualiza Cliente y crea Expediente por cada cita. Idempotente.
        Retorna un dict con estadísticas."""
        asesores_map = {u.username: u for u in User.objects.filter(profile__rol='asesor')}
        clientes_creados = clientes_existentes = 0
        expedientes_creados = expedientes_omitidos = 0
        sin_asesor = []

        for fila in filas:
            nombre = fila['nombre']
            empresa = fila['empresa']
            fecha = fila['fecha']
            hora = fila['hora'] or dt_time(9, 0)

            # 1) Cliente (get_or_create por nombre + empresa)
            cliente = Cliente.objects.filter(nombre__iexact=nombre).first()
            if cliente and (empresa and cliente.empresa != empresa):
                cliente = None  # mismo nombre pero distinta empresa → crear otro
            if not cliente:
                cliente = Cliente.objects.filter(nombre__iexact=nombre, empresa__iexact=empresa).first()
            if not cliente:
                cliente = Cliente(nombre=nombre)
                clientes_creados += 1
            else:
                clientes_existentes += 1

            cliente.empresa = empresa or cliente.empresa
            if fila['salario'] is not None:
                cliente.salario = fila['salario']
            # Mapear ubicación CLT -> oficina
            for clave, oficina in UBICACION_OFICINA.items():
                if clave in fila['ubicacion']:
                    cliente.oficina = oficina
                    break
            if not cliente.oficina:
                cliente.oficina = 'plaza_patria'  # valor por defecto
            cliente.save()

            # 2) Expediente (evitar duplicados por cliente + fecha/hora)
            if fecha:
                ya_existe = Expediente.objects.filter(
                    cliente=cliente, fecha_audiencia__date=fecha
                ).exists()
                if ya_existe:
                    expedientes_omitidos += 1
                    continue

                asesor = None
                if fila['asesor']:
                    username = _nombre_usuario(fila['asesor'])
                    asesor = asesores_map.get(username)
                    if not asesor:
                        sin_asesor.append(f"{fila['asesor']} ({nombre})")

                # Sin asesor asignable -> omitir fila (no bloquear la importación)
                if not asesor:
                    expedientes_omitidos += 1
                    continue

                notas = f'Cita CLT {fila["ubicacion"] or "-"}'
                if fila['asesor2']:
                    notas += f' | Asesor 2: {fila["asesor2"]}'
                if fila['honorario'] is not None:
                    notas += f' | Honorario est.: ${fila["honorario"]:,.2f}'
                if fila['extra']:
                    notas += f' | Extra: {fila["extra"]}'
                notas += f' | Hoja: {fila["hoja"]}'

                Expediente.objects.create(
                    cliente=cliente,
                    asesor=asesor,
                    estado='audiencia',
                    fecha_audiencia=timezone.make_aware(datetime.combine(fecha, hora)),
                    tipo_despido='injustificado',
                    notas=notas,
                )
                expedientes_creados += 1

        self.stdout.write(self.style.SUCCESS(f'  [OK] Clientes: {clientes_creados} creados, {clientes_existentes} existentes'))
        self.stdout.write(self.style.SUCCESS(
            f'  [OK] Expedientes creados: {expedientes_creados} | omitidos (duplicados): {expedientes_omitidos}'
        ))
        if sin_asesor:
            self.stdout.write(self.style.WARNING(
                f'  [!] Sin asesor asignado (usa --crear-asesores): {", ".join(sin_asesor[:10])}'
            ))

        return {
            'clientes_creados': clientes_creados,
            'clientes_existentes': clientes_existentes,
            'expedientes_creados': expedientes_creados,
            'expedientes_omitidos': expedientes_omitidos,
            'sin_asesor': sin_asesor,
        }

    def handle(self, *args, **options):
        ruta = options['archivo']
        hoja = options['hoja']
        crear_asesores = options['crear_asesores'] or options['solo_asesores']
        borrar_prueba = options['borrar_prueba']
        solo_asesores = options['solo_asesores']

        if borrar_prueba:
            self.stdout.write(self.style.WARNING('Borrando usuarios de prueba...'))
            self._borrar_usuarios_prueba()

        self.stdout.write(f'Leyendo {ruta}...')
        filas = self._leer_filas(ruta, hoja)
        if not filas:
            self.stdout.write(self.style.ERROR('No se encontraron filas para importar.'))
            return
        self.stdout.write(f'  -> {len(filas)} citas encontradas')

        if crear_asesores:
            nombres = {fila['asesor'] for fila in filas} | {fila['asesor2'] for fila in filas}
            nombres = {n for n in nombres if n and n not in ('N/A',)}
            self.stdout.write(self.style.WARNING('Creando asesores reales...'))
            self._crear_asesores(nombres)

        if solo_asesores:
            self.stdout.write(self.style.SUCCESS('Modo --solo-asesores: no se importaron citas.'))
            return

        self.stdout.write(self.style.WARNING('Importando citas...'))
        self._importar_filas(filas)
        self.stdout.write(self.style.SUCCESS('Importación completada.'))


# ─── Helpers reutilizables (usados también por la vista web) ───────────────

def _nuevo_command():
    """Instancia el comando con stdout silencioso para reutilizar su lógica."""
    return Command(stdout=StringIO())


def leer_filas(ruta, hoja=''):
    """Parsea un archivo xlsx CLT y devuelve la lista de citas (dicts)."""
    return _nuevo_command()._leer_filas(ruta, hoja)


def crear_asesores(nombres):
    """Crea (o reutiliza) usuarios asesor para los nombres dados.
    Retorna (creados, existentes)."""
    return _nuevo_command()._crear_asesores(nombres)


def importar_filas(filas):
    """Importa las citas parseadas como Cliente + Expediente.
    Retorna dict con estadísticas."""
    return _nuevo_command()._importar_filas(filas)

