# -*- coding: utf-8 -*-
"""
Genera el documento Excel detallado en español con TODAS las funciones
del Sistema de Gestión de Expedientes Laborales (Conciliación Laboral Tijuana).

Hojas:
  1. Portada          → resumen del sistema
  2. Expedientes      → funciones del módulo de expedientes
  3. Finanzas         → funciones del módulo financiero
  4. Usuarios         → cuentas, roles y permisos
  5. Extensión Chrome → funciones de la extensión
  6. Automatización   → conciliación automática, comandos y utilidades
  7. Técnica          → stack, infraestructura, despliegue

Uso:  uv run python generar_documentacion_excel.py
"""
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Estilos ──────────────────────────────────────────────────────────────
AZUL = '1F4E79'
AZUL_CLARO = 'DDEBF7'
VERDE = '375623'
VERDE_CLARO = 'E2EFDA'
GRIS = 'F2F2F2'
BORDE = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)
FUENTE_TITULO = Font(name='Calibri', size=20, bold=True, color='FFFFFF')
FUENTE_SUBTITULO = Font(name='Calibri', size=12, color='FFFFFF')
FUENTE_ENCABEZADO = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
FUENTE_CELDA = Font(name='Calibri', size=10)
FUENTE_BOLD = Font(name='Calibri', size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')
WRAP_CENTER = Alignment(wrap_text=True, vertical='center', horizontal='center')


def hoja_portada(wb):
    ws = wb.active
    ws.title = 'Portada'
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:F1')
    ws['A1'] = 'SISTEMA DE GESTIÓN DE EXPEDIENTES LABORALES'
    ws['A1'].font = FUENTE_TITULO
    ws['A1'].fill = PatternFill('solid', start_color=AZUL)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:F2')
    ws['A2'] = 'Conciliación Laboral Tijuana · Despacho Jurídico · Documento funcional del sistema'
    ws['A2'].font = FUENTE_SUBTITULO
    ws['A2'].fill = PatternFill('solid', start_color=AZUL)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    filas = [
        ('Módulos del sistema', '4 (Expedientes, Finanzas, Usuarios/Permisos, Extensión Chrome)'),
        ('Roles de usuario', 'Superadmin, Administrativo, Asesor, Abogada, Finanzas'),
        ('Núcleo', 'Gestión de casos laborales desde la primera asesoría hasta la demanda'),
        ('Tecnología', 'Python · Django 5 · PostgreSQL · Celery · Redis · Playwright · Tailwind · HTMX'),
        ('Despliegue', 'Railway (Docker) · GitHub Actions (CI)'),
        ('Documentación', 'Este archivo describe TODAS las funciones del sistema en español'),
    ]
    r = 4
    for k, v in filas:
        ws.merge_cells(f'A{r}:B{r}')
        ws[f'A{r}'] = k
        ws[f'A{r}'].font = FUENTE_BOLD
        ws[f'A{r}'].fill = PatternFill('solid', start_color=AZUL_CLARO)
        ws.merge_cells(f'C{r}:F{r}')
        ws[f'C{r}'] = v
        ws[f'C{r}'].font = FUENTE_CELDA
        ws[f'C{r}'].alignment = WRAP
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    ws[f'A{r}'] = 'Cada hoja detalla las funciones con: N° | Función | Descripción | Dónde se usa (URL/Vista) | Usuarios con acceso | Notas'
    ws[f'A{r}'].font = Font(name='Calibri', size=10, italic=True, color='555555')

    for col, ancho in zip('ABCDEF', [16, 18, 14, 18, 16, 14]):
        ws.column_dimensions[col].width = ancho
    return ws


def crear_hoja_funciones(wb, titulo, color, color_claro, encabezados, filas, anchos):
    ws = wb.create_sheet(titulo)
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(encabezados))
    ws.cell(row=1, column=1, value=titulo).font = FUENTE_TITULO
    ws.cell(row=1, column=1).fill = PatternFill('solid', start_color=color)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    for c, enc in enumerate(encabezados, 1):
        celda = ws.cell(row=2, column=c, value=enc)
        celda.font = FUENTE_ENCABEZADO
        celda.fill = PatternFill('solid', start_color=color)
        celda.alignment = WRAP_CENTER
        celda.border = BORDE
    ws.row_dimensions[2].height = 22

    for i, fila in enumerate(filas, 3):
        if i % 2 == 0:
            relleno = PatternFill('solid', start_color=GRIS)
        else:
            relleno = PatternFill('solid', start_color=color_claro)
        for c, valor in enumerate(fila, 1):
            celda = ws.cell(row=i, column=c, value=valor)
            celda.font = FUENTE_CELDA
            celda.alignment = WRAP
            celda.border = BORDE
            celda.fill = relleno
        alto = max(18, 14 * (1 + max((len(str(f)) // 60) for f in fila)))
        ws.row_dimensions[i].height = min(alto, 120)

    for c, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(c)].width = ancho
    ws.freeze_panes = 'A3'
    return ws


# ─── Hoja Expedientes ─────────────────────────────────────────────────────
ENC_EXP = ['N°', 'Función', 'Descripción', 'Dónde se usa', 'Usuarios con acceso', 'Notas']
EXPEDIENTES = [
    ('1', 'Dashboard del Asesor',
     'Panel principal del asesor: total de casos, casos activos/cerrados, próximas audiencias, mis casos recientes, alertas de prioridad alta, próximas asesorías gratuitas y avisos del admin.',
     'URL: /dashboard/asesor/', 'Asesor', 'Es la pantalla inicial tras iniciar sesión'),
    ('2', 'Dashboard Administrativo',
     'Panel del admin/superadmin: métricas globales (casos, convenios, demandas), montos reclamados/convenidos, productividad por asesor, últimos movimientos, transferencias pendientes, casos por estado, avisos y publicación de avisos obligatorios.',
     'URL: /dashboard/admin/', 'Admin, Superadmin', ''),
    ('3', 'Dashboard de la Abogada',
     'Panel exclusivo para la abogada: TODAS las demandas (objeto principal), estado de todos los clientes, calculadora laboral libre y machotes de demanda a la mano.',
     'URL: /dashboard/abogada/', 'Abogada', ''),
    ('4', 'Búsqueda global',
     'Buscador en toda la app: por número de expediente, nombre del cliente, CURP, empresa, teléfono, folio o estado.',
     'URL: /buscar/', 'Todos', ''),
    ('5', 'Listado de expedientes',
     'Lista de expedientes con filtros: búsqueda, estado, asesor, oficina, fechas y prioridad. Paginado.',
     'URL: /expedientes/', 'Todos (según rol)', 'Admin/abogada ven todos; asesor solo los suyos'),
    ('6', 'Alta de expediente',
     'Crea un expediente nuevo ligado a un cliente. Opción "Crear y Enviar a Conciliación" que lleva directo al flujo de conciliación.',
     'URL: /expedientes/nuevo/', 'Asesor, Admin, Abogada', ''),
    ('7', 'Detalle de expediente',
     'Ficha completa: datos del cliente, estado, montos, audiencia, documentos, notas, movimientos, WhatsApp, tareas de conciliación, transferencias y transiciones de estado posibles.',
     'URL: /expedientes/<pk>/', 'Todos (según rol)', ''),
    ('8', 'Edición de expediente',
     'Actualiza datos del caso. Registra movimientos de actualización y cambio de estado.',
     'URL: /expedientes/<pk>/editar/', 'Todos (según rol)', ''),
    ('9', 'Cambio de estado rápido',
     'Cambia el estado del expediente desde el detalle con transiciones válidas predefinidas (nuevo → solicitud → citatorio → audiencia → convenio, etc.).',
     'URL: /expedientes/<pk>/cambiar-estado/', 'Asesor, Admin', 'Registra movimiento automático'),
    ('10', 'Registro de resultado de audiencia',
     'Registra el resultado de la audiencia: no notificado, convenio (con monto), sin conciliación, inasistencia o reprogramada. Cambia el estado automáticamente.',
     'URL: /expedientes/<pk>/resultado-audiencia/', 'Asesor, Admin', ''),
    ('11', 'CRUD de clientes',
     'Alta, listado (con filtros), edición y detalle de clientes: datos personales, CURP, RFC, contacto, domicilio, datos laborales (salario, fechas, puesto), datos de la empresa/patrón y asesorías gratuitas.',
     'URL: /clientes/', 'Asesor, Admin', ''),
    ('12', 'Catálogo de empresas (autocompletar)',
     'Búsqueda AJAX en el catálogo de empresas importado: al elegir una empresa se rellenan automáticamente razón social, teléfono, domicilio y tipo de persona en el formulario del cliente.',
     'URL: /empresas/buscar/', 'Asesor, Admin', 'Catálogo importado desde Excel'),
    ('13', 'Documentos por expediente',
     'Sube, lista y elimina documentos del expediente (actas, identificaciones, acuses, etc.) con tipo y descripción.',
     'URL: /expedientes/<pk>/documentos/subir/', 'Todos (según rol)', 'Solo admin puede eliminar los de otros'),
    ('14', 'Notas del expediente',
     'Agrega notas internas al expediente con autor y fecha.',
     'URL: /expedientes/<pk>/notas/agregar/', 'Asesor, Admin', ''),
    ('15', 'Calendario de audiencias',
     'Calendario con vista de mes y planner semanal (lunes a viernes) de audiencias, con filtros por asesor y estado y navegación entre semanas.',
     'URL: /calendario/', 'Asesor, Admin', ''),
    ('16', 'Generador de CURP',
     'Genera una CURP a partir de nombre, apellidos, fecha de nacimiento, género y entidad (con dígito verificador).',
     'URL: /curp/', 'Todos', ''),
    ('17', 'Cálculo laboral completo',
     'Calcula todas las prestaciones desde los datos del cliente: aguinaldo, vacaciones, prima vacacional, prima de antigüedad (con tope UMA), indemnización 90 días, 20 días por año, vacaciones vencidas, horas extras, salarios devengados y días festivos. Conceptos seleccionables con checkbox.',
     'URL: /expedientes/<pk>/calculo-laboral/', 'Asesor, Admin, Abogada', 'Artículos LFT por concepto; override manual de días de vacaciones'),
    ('18', 'Simulación rápida de prestaciones',
     'Calculadora libre (sin ligar a expediente): salario + fechas + periodo de pago → estimación de prestaciones al instante.',
     'URL: /simulacion-rapida/', 'Asesor, Admin, Abogada', ''),
    ('19', 'Solicitud de conciliación',
     'Formulario oficial para iniciar la solicitud de conciliación, ligado al expediente.',
     'URL: /expedientes/<pk>/solicitud/', 'Asesor, Admin', ''),
    ('20', 'Asistente paso a paso para demanda',
     'Guía de 4 pasos: 1) datos personales del cliente, 2) información laboral, 3) empresa/patrón, 4) tipo de despido + verificación de datos críticos. Evita omitir datos.',
     'URL: /expedientes/<pk>/demanda/asistente/', 'Asesor con permiso, Admin, Abogada', ''),
    ('21', 'Editor de demanda (WYSIWYG)',
     'Editor Quill.js con la demanda generada automáticamente; permite editar el contenido antes de descargar. Selector de plantillas por tipo de despido y machotes de casos reales.',
     'URL: /expedientes/<pk>/demanda/', 'Asesor con permiso, Admin, Abogada', ''),
    ('22', 'Descarga de demanda en Word (.docx)',
     'Genera y descarga la demanda laboral en Word con la narrativa de hechos, prestaciones reclamadas y totales integrados. Verifica datos críticos antes de descargar.',
     'URL: /expedientes/<pk>/demanda/descargar/', 'Asesor con permiso, Admin, Abogada', ''),
    ('23', 'Descarga directa de demanda',
     'Versión de un clic: genera el Word sin pasar por el editor.',
     'URL: /expedientes/<pk>/demanda/directa/', 'Asesor con permiso, Admin, Abogada', ''),
    ('24', 'Machotes (biblioteca de plantillas)',
     'Catálogo de machotes legales: importar desde .docx, editar, renombrar, marcar favorito, preparar con marcadores del expediente, generar y descargar.',
     'URL: /machotes/ y /expedientes/<pk>/machotes/', 'Asesor con permiso, Admin, Abogada', 'Los marcadores {{ }} se reemplazan con datos del caso'),
    ('25', 'Guardar machote desde el editor',
     'Guarda el contenido actual del editor como machote reutilizable, convirtiendo datos específicos en marcadores.',
     'URL: /expedientes/<pk>/demanda/guardar-machote/', 'Asesor con permiso, Admin, Abogada', ''),
    ('26', 'Envío a Conciliación (3 modos)',
     'Envía la solicitud al portal de conciliación de BC con 3 opciones: (1) Automático con navegador headless, (2) Debug (navegador visible), (3) Extensión de Chrome (desde el navegador del asesor).',
     'URL: /expedientes/<pk>/conciliacion-automatica/', 'Asesor, Admin', 'Ver hoja Automatización'),
    ('27', 'Estado y screenshots de la tarea de conciliación',
     'Consulta en vivo el estado de la tarea de conciliación, ve capturas de pantalla del proceso, reintenta si falla.',
     'URL: /conciliacion/<pk>/estado/ y /procesando/', 'Asesor, Admin', ''),
    ('28', 'Subida manual del acuse (PDF)',
     'Alternativa manual: sube el PDF del acuse de conciliación, se parsea automáticamente (folio, fechas, empresa) y se confirma antes de guardar en el expediente.',
     'URL: /expedientes/<pk>/subir-conciliacion/ y /acuse-confirmar/', 'Asesor, Admin', ''),
    ('29', 'Descarga del formulario de conciliación prellenado (PDF)',
     'Genera el formato de solicitud de conciliación en PDF ya llenado con los datos del expediente.',
     'URL: /expedientes/<pk>/descargar-conciliacion/', 'Asesor, Admin', ''),
    ('30', 'WhatsApp integrado',
     'Envía mensajes de WhatsApp al cliente (deep links), plantillas predefinidas, historial por expediente y notificaciones automáticas activables.',
     'URL: /expedientes/<pk>/whatsapp/', 'Asesor, Admin', ''),
    ('31', 'Transferencia de casos entre asesores',
     'Solicita, aprueba, rechaza o cancela la transferencia de un expediente a otro asesor.',
     'URL: /transferencias/', 'Asesor (solicita), Admin (aprueba)', ''),
    ('32', 'Notificaciones internas',
     'Campana de notificaciones no leídas con marcar una/todas como leídas.',
     'URL: /notificaciones/', 'Todos', 'Context processor global'),
    ('33', 'Reportes y exportación',
     'Reportes administrativos con exportación a Excel de expedientes, y PDF del expediente (WeasyPrint).',
     'URL: /reportes/ y /reportes/excel/ y /expedientes/<pk>/pdf/', 'Admin, Superadmin', ''),
    ('34', 'Avisos obligatorios del admin',
     'El admin publica avisos/noticias con prioridad (alta/media/baja). Aparecen como modal bloqueante en TODA la app para cada usuario que no los ha leído (sin Escape ni clic fuera; solo botón "Entendido"). Fecha de vencimiento opcional: al vencer dejan de mostrarse solos.',
     'URL: /avisos/crear/ y /avisos/<pk>/marcar-leido/', 'Admin (crea), Todos (leen)', 'Seguimiento de leído por usuario'),
    ('35', 'Módulo de Ajustes',
     'Menú junto al nombre del usuario: muestra el token API personal, enlace de descarga del paquete de la extensión, instrucciones de instalación y tareas pendientes de conciliación.',
     'URL: /ajustes/', 'Todos', ''),
    ('36', 'Importación de citas CLT.xlsx',
     'Importa citas de conciliación desde el archivo Excel CLT.xlsx de forma masiva.',
     'URL: /clt/importar/', 'Admin, Superadmin', ''),
    ('37', 'Importación de machotes (.docx)',
     'Sube archivos Word como machotes reutilizables con detección de marcadores.',
     'URL: /machotes/importar/', 'Admin, Superadmin', ''),
]
ANCHOS_EXP = [5, 26, 60, 30, 24, 32]

# ─── Hoja Finanzas ────────────────────────────────────────────────────────
ENC_FIN = ['N°', 'Función', 'Descripción', 'Dónde se usa', 'Usuarios con acceso', 'Notas']
FINANZAS = [
    ('1', 'Dashboard financiero',
     'Panel con indicadores clave: flujo de caja mensual, ingresos/egresos, utilidades y acceso rápido a los módulos financieros.',
     'URL: /dashboard/financiero/', 'Finanzas, Admin, Superadmin', ''),
    ('2', 'Exportar dashboard financiero a Excel',
     'Descarga el resumen financiero del dashboard en formato Excel.',
     'URL: /dashboard/financiero/exportar-excel/', 'Finanzas, Admin', ''),
    ('3', 'Flujo mensual (API)',
     'API interna que alimenta la gráfica de flujo de caja mensual.',
     'URL: /api/flujo-mensual/', 'Finanzas, Admin', ''),
    ('4', 'Movimientos de caja (CRUD completo)',
     'Registra, lista (con filtros), edita y elimina movimientos de caja: ingresos, egresos, categorías y montos.',
     'URL: /caja/', 'Finanzas, Admin', ''),
    ('5', 'Socios del despacho',
     'Administra los socios: alta, detalle, edición. Base para convenios, préstamos y distribución de utilidades.',
     'URL: /socios/', 'Finanzas, Admin', ''),
    ('6', 'Semanas de trabajo',
     'Registra las semanas laborales (base para cálculo de comisiones y pagos).',
     'URL: /semanas/', 'Finanzas, Admin', ''),
    ('7', 'Préstamos entre socios',
     'Registra y administra préstamos entre socios del despacho.',
     'URL: /prestamos/', 'Finanzas, Admin', ''),
    ('8', 'Convenios (agreements)',
     'Gestión de convenios: alta, detalle, edición. Incluye honorarios vinculados.',
     'URL: /convenios/', 'Finanzas, Admin', ''),
    ('9', 'Honorarios',
     'Registra honorarios por convenio (alta y edición).',
     'URL: /honorarios/', 'Finanzas, Admin', ''),
    ('10', 'Distribución de utilidades',
     'Calcula y registra la distribución de utilidades entre socios: alta, detalle, edición y confirmación.',
     'URL: /distribuciones/', 'Finanzas, Admin', ''),
    ('11', 'Reporte de convenios',
     'Reporte de convenios con exportación a Excel.',
     'URL: /reportes/convenios/ y /exportar-excel/', 'Finanzas, Admin', ''),
    ('12', 'Modelos financieros internos',
     'Oficinas, pagos de liquidación, gastos, comisiones, empleados, nóminas, resumen de utilidades por socio: respaldan los cálculos del módulo.',
     'Interno (modelos de datos)', 'Finanzas, Admin', ''),
]
ANCHOS_FIN = [5, 26, 60, 32, 24, 26]

# ─── Hoja Usuarios ────────────────────────────────────────────────────────
ENC_USR = ['N°', 'Función', 'Descripción', 'Dónde se usa', 'Usuarios con acceso', 'Notas']
USUARIOS = [
    ('1', 'Roles del sistema',
     'Cinco roles con distintos niveles de acceso: Superadmin, Administrativo (admin), Asesor, Abogada y Finanzas. El perfil de cada usuario guarda su rol y permisos.',
     'Modelo UserProfile', 'Todos', ''),
    ('2', 'Inicio de sesión personalizado',
     'Login que redirige a cada usuario según su rol (asesor → dashboard asesor, abogada → dashboard abogada, finanzas → dashboard financiero, admin/superadmin → dashboard admin).',
     'URL: /accounts/login/', 'Todos', ''),
    ('3', 'Cierre de sesión',
     'Logout con redirección al login.',
     'URL: /accounts/logout/', 'Todos', ''),
    ('4', 'Recuperación de contraseña',
     'Flujo completo de restablecimiento: solicitud por email, correo con enlace, confirmación y nueva contraseña.',
     'URL: /accounts/password-reset/', 'Todos', ''),
    ('5', 'Dashboard del Superadmin',
     'Panel exclusivo del superadmin para la administración avanzada del sistema.',
     'URL: /accounts/superadmin/', 'Superadmin', ''),
    ('6', 'Matriz de permisos',
     'Matriz que muestra/administra permisos por rol, con exportación a Excel.',
     'URL: /accounts/superadmin/matriz-permisos/', 'Superadmin', ''),
    ('7', 'Carga de datos demo',
     'Carga datos de demostración para probar el sistema.',
     'URL: /accounts/superadmin/cargar-datos-demo/', 'Superadmin', ''),
    ('8', 'Auditoría de permisos',
     'Registro de auditoría de cambios de permisos (PermisoAuditLog).',
     'Modelo PermisoAuditLog', 'Superadmin', ''),
    ('9', 'Token API por usuario (extensión)',
     'Cada usuario tiene un token API único (auto-generado) para autenticar la extensión de Chrome con la app. Se puede regenerar desde Ajustes o la página de configuración de la extensión.',
     'URL: /extension/config/ y /extension/regenerar-token/', 'Todos', 'Se muestra en Ajustes'),
]
ANCHOS_USR = [5, 26, 60, 36, 24, 26]

# ─── Hoja Extensión Chrome ────────────────────────────────────────────────
ENC_EXT = ['N°', 'Función', 'Descripción', 'Dónde se usa', 'Usuarios con acceso', 'Notas']
EXTENSION = [
    ('1', 'Instalación guiada',
     'El asesor descarga el paquete .zip desde Ajustes, lo descomprime y lo carga en chrome://extensions (Modo desarrollador → Cargar descomprimida). Guía paso a paso en la página de configuración y en LEEME.md.',
     'URL: /ajustes/ y /extension/descargar/', 'Todos', 'Instalación única por computadora'),
    ('2', 'Configuración de la extensión',
     'La página de opciones de la extensión guarda la URL de la app y el token personal del asesor (chrome.storage).',
     'Archivos: options.html, options.js', 'Todos', ''),
    ('3', 'Popup con tareas pendientes',
     'Al hacer clic en el ícono de la extensión se listan las tareas de conciliación pendientes (modo extensión) con el botón "Llenar".',
     'Archivos: popup.html, popup.js', 'Asesor', ''),
    ('4', 'Llenado automático del portal del gobierno',
     'El content script llena el portal app.conciliacionbc.gob.mx en el navegador REAL del asesor: datos del solicitante (nombre mexicano compuesto separado correctamente), CURP real, domicilio, datos laborales, empresa citada, jornada y narrativa de hechos.',
     'Archivo: content.js', 'Asesor', 'El asesor ve todo en vivo y resuelve CAPTCHA'),
    ('5', 'Detección de folio y acuse',
     'La extensión detecta el folio del acuse tras el envío, descarga el PDF del acuse y captura screenshot.',
     'Archivo: content.js', 'Asesor', ''),
    ('6', 'Reporte automático a la app',
     'El service worker envía el folio, el PDF del acuse y los screenshots a la app vía API con token. La app guarda el folio en el expediente, el acuse como documento y marca la tarea completada.',
     'API: /api/extension/tareas/<id>/reportar/', 'Asesor', 'Auth: Authorization: Token <token>'),
    ('7', 'API de tareas para la extensión',
     'GET /api/extension/tareas/ devuelve las tareas pendientes con todos los datos del cliente/empresa y la narrativa para llenar el portal.',
     'API: /api/extension/tareas/', 'Asesor', 'Filtra solo modo=extension'),
    ('8', 'Estado del proceso en vivo',
     'Panel flotante en el portal que muestra el paso actual del llenado (aviso → industria → fecha → solicitante → empresa → descripción → envío).',
     'Archivo: content.js', 'Asesor', ''),
]
ANCHOS_EXT = [5, 26, 60, 34, 22, 30]

# ─── Hoja Automatización ──────────────────────────────────────────────────
ENC_AUT = ['N°', 'Función', 'Descripción', 'Dónde se usa', 'Usuarios con acceso', 'Notas']
AUTOMATIZACION = [
    ('1', 'Conciliación automática (headless)',
     'Automatización con Playwright que llena el portal de conciliación de BC en navegador headless del servidor: 8 fases completas. Ejecución asíncrona con threading (o Celery si hay Redis).',
     'expedientes/conciliacion_automation.py', 'Asesor, Admin', 'Tarea con estado, capturas y reintentos'),
    ('2', 'Validador de CURP real',
     'Valida la CURP del cliente contra el formato oficial RENAPO y exige CURP real (rechaza placeholders XAXX010101000, inventados o sintéticos) antes de enviar al portal.',
     'expedientes/conciliacion_automation.py', 'Asesor, Admin', ''),
    ('3', 'Parser del acuse PDF',
     'Parsea el acuse de conciliación (PDF) para extraer folio, fechas y datos automáticamente al subirlo manualmente.',
     'expedientes/acuse_parser.py', 'Asesor, Admin', ''),
    ('4', 'Comandos programados',
     'Comandos para tareas periódicas: enviar recordatorios, enviar solicitudes de conciliación, enviar WhatsApp automático, crear usuarios de prueba, sembrar datos.',
     'expedientes/management/commands/', 'Admin (cron/Celery beat)', ''),
    ('5', 'Importación de datos',
     'Importa CLT.xlsx (citas), catálogo de empresas y machotes desde archivos Excel/Word.',
     'Comandos: importar_clt, importar_empresas, importar_machotes', 'Admin, Superadmin', ''),
    ('6', 'Migración de SQLite a PostgreSQL',
     'Herramienta para migrar la base de datos local a PostgreSQL (producción).',
     'Comando: migrate_sqlite_to_pg', 'Superadmin', ''),
    ('7', 'Semillas de datos de prueba',
     'seed_clientes_prueba genera 9 casos de prueba con datos completos por tipo de despido y cálculos laborales (verificados por el CI).',
     'Comando: seed_clientes_prueba', 'Desarrollo/CI', ''),
    ('8', 'Verificación de cálculos (CI)',
     'Script que verifica que los conceptos de cada demanda sean coherentes con el tipo de despido (renuncia voluntaria no reclama indemnización) y que los totales coincidan con el cálculo guardado. Corre en el CI.',
     'verify_demandas_calculos.py', 'CI (GitHub Actions)', ''),
]
ANCHOS_AUT = [5, 26, 60, 34, 24, 30]

# ─── Hoja Técnica ─────────────────────────────────────────────────────────
ENC_TEC = ['Aspecto', 'Detalle', 'Notas']
TECNICA = [
    ('Stack principal', 'Python 3.13 · Django 5 · PostgreSQL (producción) / SQLite (desarrollo)', ''),
    ('Frontend', 'Tailwind CSS (CDN) + HTMX + Quill.js (editor WYSIWYG)', 'Templates con estilos propios'),
    ('Generación de documentos', 'python-docx (Word), WeasyPrint (PDF), openpyxl (Excel)', ''),
    ('Automatización del portal', 'Playwright (headless y debug) + Extensión Chrome MV3', ''),
    ('Tareas asíncronas', 'Celery + Redis si están disponibles; fallback con threading', 'Ver entrypoint.sh'),
    ('Estáticos en producción', 'Whitenoise (CompressedManifestStaticFilesStorage)', 'collectstatic en despliegue'),
    ('CI/CD', 'GitHub Actions: check, 86 tests, verificación de cálculos', 'Workflow: .github/workflows/ci.yml'),
    ('Despliegue', 'Railway con Dockerfile (gunicorn + entrypoint.sh con migraciones y seed)', 'railway.json'),
    ('Extension Chrome', 'Manifest V3 con popup, options, content script y service worker', 'Carpeta extension/'),
    ('Seguridad', 'Auth de Django, roles por perfil, CSRF, token API para la extensión, validación de CURP', ''),
    ('Calidad', '86 tests automatizados (expedientes) que pasan en CI (Linux/UTF-8)', 'settings_test'),
    ('Módulos del sistema', 'expedientes (núcleo) · finanzas · accounts · core/laboral (calculadoras)', ''),
]
ANCHOS_TEC = [30, 70, 40]


def main():
    wb = Workbook()
    hoja_portada(wb)
    crear_hoja_funciones(wb, 'EXPEDIENTES (núcleo)', AZUL, AZUL_CLARO, ENC_EXP, EXPEDIENTES, ANCHOS_EXP)
    crear_hoja_funciones(wb, 'FINANZAS', VERDE, VERDE_CLARO, ENC_FIN, FINANZAS, ANCHOS_FIN)
    crear_hoja_funciones(wb, 'USUARIOS Y PERMISOS', '7030A0', 'E4DFEC', ENC_USR, USUARIOS, ANCHOS_USR)
    crear_hoja_funciones(wb, 'EXTENSIÓN DE CHROME', 'BF6000', 'FDE9D9', ENC_EXT, EXTENSION, ANCHOS_EXT)
    crear_hoja_funciones(wb, 'AUTOMATIZACIÓN', '8B3A3A', 'F5E1E1', ENC_AUT, AUTOMATIZACION, ANCHOS_AUT)
    crear_hoja_funciones(wb, 'TÉCNICA (stack)', '404040', 'EDEDED', ENC_TEC, TECNICA, ANCHOS_TEC)

    nombre = 'Documentacion_Sistema_Conciliacion_Laboral.xlsx'
    wb.save(nombre)
    print(f'✅ Documento generado: {nombre}')
    print(f'   Hojas: {wb.sheetnames}')


if __name__ == '__main__':
    main()
